import requests, bs4 #imports requests and bs4
from openpyxl import Workbook #imports workbook from openpyxl

res = requests.get('https://eservices.minnstate.edu/registration/search/advancedSubmit.html?campusid=305&searchrcid=0305&searchcampusid=305&yrtr=20205&subject=ITEC&courseNumber=&courseId=&openValue=OPEN_PLUS_WAITLIST&delivery=ALL&showAdvanced=&starttime=&endtime=&mntransfer=&credittype=ALL&credits=&instructor=&keyword=&begindate=&site=&resultNumber=250') #sets requests equal to requesting info from a url
eservices_file = res.content #sets eservices_file to res.contents

workbook = Workbook() #sets workbook to Workbook()

worksheet = workbook.active #sets worksheet to an active workbook

worksheet.title = 'FINAL' #sets worksheet title to final

worksheet.cell(1, 1, 'ID') #first cell in column 1 is called ID
worksheet.cell(1, 2, 'Course #') #first cell in column 2 is called Course #
worksheet.cell(1, 3, 'Course Title') #first cell in column 3 is called Course title
worksheet.cell(1, 4, 'Day') #first cell in column 4 is called Day
worksheet.cell(1, 5, 'Time') #first cell in column 5 is called Time
worksheet.cell(1, 6, 'Credit') #first cell in column 6 is called Credit
worksheet.cell(1, 7, 'Instructor') #first cell in column 7 is called Instuctor


    



eservices_soup = bs4.BeautifulSoup(eservices_file, features="lxml") #sets eservices_soup to the eservices_file
table = eservices_soup.find_all('tbody')[1] #sets table to find tbody from eservices_soup
rows = table.find_all('tr') #finds all tr in html document
ssrow = 2 #sets ssrow equal to 2
for row in rows: # for loop for row in rows
    td = row.find_all('td') #finds all td in html document
    data = [] #data as a list
    for index, stuff in enumerate(td): #for loop to enumerate td from html document
        data.append(stuff.text.strip()) #adds stuff to the data list
    worksheet.cell(ssrow, 1, data[1]) #calls course ID from list and writes it to the excel document
    worksheet.cell(ssrow, 2, data[3]) #calls course number from list and writes it to the excel document 
    worksheet.cell(ssrow, 3, data[5]) #calls course title from list and writes it to the excel document
    worksheet.cell(ssrow, 4, data[7]) #calls day from list and writes it to the excel document
    worksheet.cell(ssrow, 5, data[8]) #calls time from list and writes it to the excel document 
    worksheet.cell(ssrow, 6, data[9]) #calls credit from list and writes it to the excel document
    worksheet.cell(ssrow, 7, data[11]) #calls instuctor from list and writes it to the excel document
    
    ssrow += 1 #adds 1 to ssrow after every for loop iteration

workbook.save('final.xlsx') #saves file

