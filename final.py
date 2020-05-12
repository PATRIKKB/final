import requests, bs4
from openpyxl import Workbook

url = 'https://eservices.minnstate.edu/registration/search/advancedSubmit.html?campusid=305&searchrcid=0305&searchcampusid=305&yrtr=20205&subject=ITEC&courseNumber=&courseId=&openValue=OPEN_PLUS_WAITLIST&delivery=ALL&showAdvanced=&starttime=&endtime=&mntransfer=&credittype=ALL&credits=&instructor=&keyword=&begindate=&site=&resultNumber=250'

ID = ['narg', 'wonkle', 'boopra']
course_number = ['fart fart boop', 'fartlord cheese']
course_title = ['pooping on a toilet']
day = ['wens', 'tues']
time = ['4pm', '5pm']
credit = ['500']
instructor = ['cow', 'alargeblackdor']


workbook = Workbook()

worksheet = workbook.active

worksheet.title = 'FINAL'

worksheet.cell(1, 1, 'ID')
worksheet.cell(1, 2, 'Course #')
worksheet.cell(1, 3, 'Course Title')
worksheet.cell(1, 4, 'Day')
worksheet.cell(1, 5, 'Time')
worksheet.cell(1, 6, 'Credit')
worksheet.cell(1, 7, 'Instructor')

for index, contents in enumerate(ID):
    worksheet.cell(index + 2, 1, contents)

for index, contents in enumerate(course_number):
    worksheet.cell(index + 2, 2, contents)

for index, contents in enumerate(course_title):
    worksheet.cell(index + 2, 3, contents)

for index, contents in enumerate(day):
    worksheet.cell(index + 2, 4, contents)

for index, contents in enumerate(time):
    worksheet.cell(index + 2, 5, contents)

for index, contents in enumerate(credit):
    worksheet.cell(index + 2, 6, contents)

for index, contents in enumerate(instructor):
    worksheet.cell(index + 2, 7, contents)
    
workbook.save('final.xlsx')
